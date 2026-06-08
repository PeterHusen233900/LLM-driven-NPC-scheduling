"""
generate_qualitative_sample.py
-------------------------------
Generates a stratified, diversity-weighted 20% qualitative sample (640 rows)
from the full NPC schedule evaluation dataset (3200 schedules across 4 models,
5 scenarios, 2 rule conditions, 8 NPCs, 10 runs).
 
Sampling strategy (two-stage):
  Stage 1 — Guarantee: minimum MIN_PER_NPC schedules per NPC per model
  Stage 2 — Fill:      remaining slots filled using a diversity score as weight,
                       ensuring representation across scenarios and complexity tiers
 
Diversity score combines:
  - Number of unique action types in the schedule
  - Number of non-WALK_AROUND steps (capped)
  - Total step count (capped)
 
Tiers (derived from diversity score):
  degenerate  score <= 2   (single or trivial action)
  low         score  3–4
  medium      score  5–9
  high        score >= 10
 
Output: outputs/qualitative_sample_640.xlsx with three sheets:
  - Scoring              : 640 rows — SampleID, NPC, scenario, Rule Check, Goal,
                           Steps, P1_Personality, P2_Goal_Coherence,
                           P3_Threat_Awareness (High Threat only), Notes
  - Model Key (reveal after) : maps SampleID → model, NPC, scenario, Run,
                               diversity_score
  - Sample Composition   : breakdown by tier, model, and scenario
 
Usage:
  python generate_qualitative_sample.py
 
Edit EVAL_FILES below to point at your actual timestamped xlsx files.
"""


import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

RANDOM_SEED    = 42
MIN_PER_NPC    = 4       # guaranteed minimum samples per NPC per model
TOTAL_PER_MODEL = 160    # 80 full + 80 minimal = 160 per model, 640 total
SLOTS_PER_CONDITION = 80 # samples per model per rule condition

SCENARIOS = ["Baseline", "Village Safe", "Keys Available", "Safe + Keys", "High Threat"]

EVAL_FILES = {
    "GPT_full":    "../outputs/npc_evaluation_20260531_111736_GPT-OSS-120B.xlsx",
    "GPT_minimal": "../outputs/npc_evaluation_20260531_142340_GPT-OSS-120B_minimal.xlsx",
    "Qwen122_full":    "../outputs/npc_evaluation_20260531_120357_Qwen3.5-122B.xlsx",
    "Qwen122_minimal": "../outputs/npc_evaluation_20260531_142340_Qwen3.5-122B_minimal.xlsx",
    "Qwen27_full":    "../outputs/npc_evaluation_20260531_120357_Qwen3.6-27B.xlsx",
    "Qwen27_minimal": "../outputs/npc_evaluation_20260531_142340_Qwen3.6-27B_minimal.xlsx",
    "Llama_full":    "../outputs/npc_evaluation_20260531_120357_Llama3.3-70B.xlsx",
    "Llama_minimal": "../outputs/npc_evaluation_20260531_142340_Llama3.3-70B_minimal.xlsx",
}

OUTPUT_PATH = "../scoring/qualitative_sample_640.xlsx"

# ── Diversity score & tier ─────────────────────────────────────────────────────
 
def diversity_score(steps_str: str) -> int:
    """
    Score a schedule by how diverse and substantive its actions are.
    Combines: unique action types + non-WALK_AROUND steps + capped total steps.
    """
    if not isinstance(steps_str, str):
        return 0
    lines = [re.sub(r"^\d+\.\s*", "", l.strip())
             for l in steps_str.strip().split("\n") if l.strip()]
    if not lines:
        return 0
    actions = [m.group(1) for l in lines
               for m in [re.match(r"([A-Z_]+)\(", l)] if m]
    if not actions:
        return 0
    unique_types   = len(set(actions))
    non_walk_count = sum(1 for a in actions if a != "WALK_AROUND")
    total_capped   = min(len(actions), 15)
    return unique_types + non_walk_count + total_capped
 
 
def compute_tier(score: int) -> str:
    if score <= 2:
        return "Degenerate"
    elif score <= 4:
        return "Low"
    elif score <= 9:
        return "Medium"
    else:
        return "High"
 
 
# ── Load all evaluation data ───────────────────────────────────────────────────
 
def load_all_schedules() -> pd.DataFrame:
    all_rows = []
    for model_key, filepath in EVAL_FILES.items():
        try:
            xl = pd.read_excel(filepath, sheet_name=None)
        except FileNotFoundError:
            print(f"  WARNING: file not found — {filepath}")
            continue
        for scenario in SCENARIOS:
            sheet_key = f"{scenario} Schedules"
            if sheet_key not in xl:
                continue
            df = xl[sheet_key].copy()
            df = df[df["Status"] == "OK"].copy()
            df["model"]    = model_key
            df["scenario"] = scenario
            all_rows.append(df)
    if not all_rows:
        raise RuntimeError("No evaluation data loaded. Check EVAL_FILES paths.")
    combined = pd.concat(all_rows, ignore_index=True)
    print(f"Loaded {len(combined)} valid schedules.")
    return combined
 
 
# ── Stratified weighted sampling ──────────────────────────────────────────────
 
def sample_one_condition(df: pd.DataFrame, model_key: str,
                         n_slots: int, seed: int) -> pd.DataFrame:
    """Two-stage stratified weighted sample for one model × rule-condition slice."""
    subset = df[df["model"] == model_key].copy()
    if subset.empty:
        return pd.DataFrame()
 
    subset["diversity_score"] = subset["Steps"].apply(diversity_score)
    subset["tier"] = subset["diversity_score"].apply(compute_tier)
 
    # Stage 1: guaranteed minimum per NPC
    guaranteed, used_idx = [], set()
    for npc, group in subset.groupby("NPC"):
        sampled = group.sample(n=min(MIN_PER_NPC, len(group)), random_state=seed)
        guaranteed.append(sampled)
        used_idx.update(sampled.index)
 
    guaranteed_df = pd.concat(guaranteed, ignore_index=True) if guaranteed else pd.DataFrame()
    pool = subset[~subset.index.isin(used_idx)].copy()
 
    # Stage 2: diversity-weighted fill
    n_remaining = max(0, n_slots - len(guaranteed_df))
    if n_remaining > 0 and not pool.empty:
        weights = pool["diversity_score"].clip(lower=0.1)
        fill_df = pool.sample(n=min(n_remaining, len(pool)),
                              weights=weights, random_state=seed, replace=False)
    else:
        fill_df = pd.DataFrame()
 
    return pd.concat([guaranteed_df, fill_df], ignore_index=True).head(n_slots)
 
 
def build_sample(df: pd.DataFrame) -> pd.DataFrame:
    all_samples = []
    counter = 1
    for model_key in EVAL_FILES:
        sampled = sample_one_condition(df, model_key, SLOTS_PER_CONDITION, RANDOM_SEED)
        sampled = sampled.copy()
        sampled["SampleID"] = [f"S{counter + i:03d}" for i in range(len(sampled))]
        all_samples.append(sampled)
        counter += len(sampled)
        print(f"  {model_key}: {len(sampled)} samples")
    return pd.concat(all_samples, ignore_index=True)
 
 
# ── Excel output ───────────────────────────────────────────────────────────────
 
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BOLD_FONT    = Font(bold=True, name="Arial", size=10)
DEFAULT_FONT = Font(name="Arial", size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
TOP_ALIGN    = Alignment(vertical="top")
 
 
def set_header(ws, row: int, cols: list) -> None:
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
 
 
def write_scoring_sheet(ws, sample: pd.DataFrame) -> None:
    cols = ["SampleID", "NPC", "scenario", "Rule Check", "Goal", "Steps",
            "P1_Personality", "P2_Goal_Coherence", "P3_Threat_Awareness", "Notes"]
    set_header(ws, 1, cols)
 
    for r, (_, row) in enumerate(sample.iterrows(), 2):
        ws.cell(r, 1, row["SampleID"]).font        = DEFAULT_FONT
        ws.cell(r, 2, row.get("NPC", "")).font     = DEFAULT_FONT
        ws.cell(r, 3, row.get("scenario", "")).font = DEFAULT_FONT
 
        rule_check = row.get("Rule Check", row.get("Validity", ""))
        ws.cell(r, 4, str(rule_check)).font = DEFAULT_FONT
 
        goal = row.get("Goal", row.get("goal", ""))
        ws.cell(r, 5, str(goal) if pd.notna(goal) else "").font = DEFAULT_FONT
 
        steps_cell = ws.cell(r, 6, str(row.get("Steps", "")))
        steps_cell.alignment = WRAP_ALIGN
        steps_cell.font      = DEFAULT_FONT
 
        # P1, P2 left blank for scoring; P3 only relevant for High Threat
        ws.cell(r, 7, None)  # P1_Personality
        ws.cell(r, 8, None)  # P2_Goal_Coherence
        p3 = None if row.get("scenario") != "High Threat" else None
        ws.cell(r, 9, p3)    # P3_Threat_Awareness
        ws.cell(r, 10, None) # Notes
 
    col_widths = [8, 12, 14, 12, 45, 70, 14, 18, 20, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(sample) + 1}"
 
 
def write_model_key_sheet(ws, sample: pd.DataFrame) -> None:
    cols = ["SampleID", "model", "NPC", "scenario", "Run", "diversity_score"]
    set_header(ws, 1, cols)
 
    run_col = "Run" if "Run" in sample.columns else None
 
    for r, (_, row) in enumerate(sample.iterrows(), 2):
        ws.cell(r, 1, row["SampleID"]).font             = DEFAULT_FONT
        ws.cell(r, 2, row.get("model", "")).font        = DEFAULT_FONT
        ws.cell(r, 3, row.get("NPC", "")).font          = DEFAULT_FONT
        ws.cell(r, 4, row.get("scenario", "")).font     = DEFAULT_FONT
        run_val = int(row[run_col]) if run_col and pd.notna(row.get(run_col)) else None
        ws.cell(r, 5, run_val).font                     = DEFAULT_FONT
        score = row.get("diversity_score", 0)
        ws.cell(r, 6, int(score) if pd.notna(score) else 0).font = DEFAULT_FONT
 
    col_widths = [8, 18, 12, 14, 6, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    ws.freeze_panes = "A2"
 
 
def write_composition_sheet(ws, sample: pd.DataFrame) -> None:
    rows = [
        ("Metric", "Value"),
        ("Total schedules in population", 3200),
        ("Sample size", len(sample)),
        ("Sample %", f"{len(sample) / 3200 * 100:.1f}%"),
        ("Random seed", RANDOM_SEED),
        ("Min per NPC per model", MIN_PER_NPC),
        ("Weighted slots per model", SLOTS_PER_CONDITION - MIN_PER_NPC * 8),
        (None, None),
        ("Tier", "Count"),
    ]
    for tier, count in sample["tier"].value_counts().sort_index().items():
        rows.append((tier, int(count)))
 
    rows += [(None, None), ("Model", "Count")]
    for model, count in sorted(sample["model"].value_counts().items()):
        rows.append((model, int(count)))
 
    rows += [(None, None), ("Scenario", "Count")]
    for sc, count in sample["scenario"].value_counts().sort_index().items():
        rows.append((sc, int(count)))
 
    headers = {"Metric", "Tier", "Model", "Scenario"}
    for r, (label, val) in enumerate(rows, 1):
        cell_a = ws.cell(r, 1, label)
        cell_b = ws.cell(r, 2, val)
        if label in headers or r == 1:
            cell_a.font = BOLD_FONT
            cell_b.font = BOLD_FONT
        else:
            cell_a.font = DEFAULT_FONT
            cell_b.font = DEFAULT_FONT
 
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
 
 
def save_workbook(sample: pd.DataFrame, path: str) -> None:
    wb = Workbook()
 
    ws_score = wb.active
    ws_score.title = "Scoring"
    write_scoring_sheet(ws_score, sample)
 
    ws_key = wb.create_sheet("Model Key (reveal after)")
    write_model_key_sheet(ws_key, sample)
 
    ws_comp = wb.create_sheet("Sample Composition")
    write_composition_sheet(ws_comp, sample)
 
    wb.save(path)
    print(f"\nSaved: {path}")
 
 
# ── Main ───────────────────────────────────────────────────────────────────────
 
def main() -> None:
    print("=" * 60)
    print("Qualitative Sample Generator")
    print("=" * 60)
 
    df = load_all_schedules()
 
    print(f"\nBuilding sample (target: {SLOTS_PER_CONDITION * len(EVAL_FILES)} rows)...")
    sample = build_sample(df)
 
    print(f"\nTotal samples : {len(sample)}")
    print("Tier breakdown:", sample["tier"].value_counts().to_dict())
    print("Model counts  :", sample["model"].value_counts().to_dict())
 
    save_workbook(sample, OUTPUT_PATH)
 
 
if __name__ == "__main__":
    main()
 
