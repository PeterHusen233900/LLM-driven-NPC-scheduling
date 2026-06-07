"""
run.py — NPC Schedule Evaluation Runner
----------------------------------------
Loads all game data from external JSON files, then sends generate requests
to the npc_generation API for all NPCs. Repeats for NUM_RUNS total runs
across all scenario world states, and saves results to a single xlsx file
per model under /outputs — one Schedules sheet and one Summary sheet per scenario.

The workbook is saved after every NPC completes, so progress is never lost
if the script is interrupted mid-run.

Expected file layout:
    inputs/
        ..world_state.json                            # baseline world state
        scenario_1_village_safe.json                # scenario variants
        scenario_2_keys_available.json
        scenario_3_village_safe_keys_available.json
        scenario_4_high_threat.json
        world_rules.json
        action_library.json
        characters.json
    outputs/                                        # xlsx written here

Usage:
    pip install requests openpyxl
    python run.py
"""

import os
import sys
import json
import time
from datetime import datetime

import requests
from validator import validate_schedule as rule_validate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

API_BASE    = "http://localhost:8001"
NUM_RUNS    = 10
DATA_DIR    = "../inputs"
OUTPUT_DIR  = "outputs"
TEMPERATURE = 0.7
MAX_TOKENS  = 4096

# Models to evaluate — each gets its own output file
MODELS = [
    "GPT-OSS-120B",
    "Qwen3.5-122B",
    "Qwen3.6-27B",
    "Llama3.3-70B",
]

# Scenarios to run in order — (sheet_label, world_state_filename)
SCENARIOS = [
    ("Baseline",          "world_state.json"),
    ("Village Safe",      "scenario_1_village_safe.json"),
    ("Keys Available",    "scenario_2_keys_available.json"),
    ("Safe + Keys",       "scenario_3_village_safe_keys_available.json"),
    ("High Threat",       "scenario_4_high_threat.json"),
]

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Load game data from files
# ---------------------------------------------------------------------------

def load_json(filename: str) -> dict | list:
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        print(f"ERROR: Required data file not found: {path}")
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_all(world_state_file: str) -> tuple[dict, dict, dict, list]:
    world_state    = load_json(world_state_file)
    world_rules    = load_json("world_rules.json")
    action_library = load_json("action_library.json")
    characters     = load_json("characters.json")
    return world_state, world_rules, action_library, characters

# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def generate_schedule(npc_name: str, npc_data: dict,
                      world_state: dict, world_rules: dict,
                      action_library: dict, model: str) -> dict:
    payload = {
        "npc_name":              npc_name,
        "character_description": npc_data,
        "world_state":           world_state,
        "world_rules":           world_rules,
        "action_library":        action_library,
        "temperature":           TEMPERATURE,
        "max_tokens":            MAX_TOKENS,
        "model":                 model,
    }
    t0 = time.time()
    resp = requests.post(f"{API_BASE}/generate", json=payload, timeout=120)
    latency = round(time.time() - t0, 2)
    resp.raise_for_status()
    data = resp.json()

    usage = data.get("usage", {})
    return {
        "schedule":      data["schedule"],
        "latency_s":     latency,
        "input_tokens":  usage.get("input_tokens",  None),
        "output_tokens": usage.get("output_tokens", None),
    }

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F1F1F", end_color="1F1F1F")
ALT_FILL    = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_WRAP   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_COLS = {1, 2, 5, 6, 7, 8}


def _header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER


def _style_row(ws, row_idx: int, col_count: int, alt: bool):
    fill = ALT_FILL if alt else None
    for col in range(1, col_count + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font      = BODY_FONT
        c.border    = THIN_BORDER
        c.alignment = CENTER if col in CENTER_COLS else LEFT_WRAP
        if fill:
            c.fill = fill


def _steps_text(steps: list[str]) -> str:
    return "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps)) if steps else "—"


def _write_schedules_sheet(ws, results: list[dict]):
    headers = [
        "Run", "NPC", "Goal", "Steps",
        "Gen Latency (s)", "Input Tokens", "Output Tokens",
        "Status", "Rule Check",
        "Root Violations", "Cascading Violations", "Violation Detail",
    ]
    _header(ws, headers)
    col_widths = [6, 18, 40, 72, 16, 14, 15, 12, 12, 8, 8, 60]
    for col, width in zip("ABCDEFGHIJKL", col_widths):
        ws.column_dimensions[col].width = width

    for i, e in enumerate(results, start=2):
        ws.cell(row=i, column=1,  value=e["run"])
        ws.cell(row=i, column=2,  value=e["npc"])
        ws.cell(row=i, column=3,  value=e["goal"])
        ws.cell(row=i, column=4,  value=_steps_text(e["steps"]))
        ws.cell(row=i, column=5,  value=e["gen_latency_s"])
        ws.cell(row=i, column=6,  value=e.get("input_tokens"))
        ws.cell(row=i, column=7,  value=e.get("output_tokens"))
        ws.cell(row=i, column=8,  value=e["gen_status"])
        ws.cell(row=i, column=9,  value=("VALID" if e.get("rule_valid") else
                                          ("INVALID" if e.get("rule_valid") is False else "—")))
        ws.cell(row=i, column=10, value=e.get("rule_root", ""))
        ws.cell(row=i, column=11, value=e.get("rule_cascading", ""))
        ws.cell(row=i, column=12, value="\n".join(e.get("rule_violations", [])) or "—")
        _style_row(ws, i, 12, i % 2 == 0)
        ws.row_dimensions[i].height = max(15, len(e["steps"]) * 14) if e["steps"] else 15


def _write_summary_sheet(ws, results: list[dict], npc_names: list[str]):
    _header(ws, [
        "NPC", "Gen Success Rate", "Avg Gen Latency (s)", "Avg Steps",
        "Avg Input Tokens", "Avg Output Tokens", "Avg Total Tokens",
        "Rule Valid Rate", "Avg Root Violations", "Avg Cascading Violations",
    ])
    col_widths = [22, 18, 20, 12, 18, 18, 18, 16, 20, 24]
    for col, width in zip("ABCDEFGHIJ", col_widths):
        ws.column_dimensions[col].width = width

    def avg(lst):
        clean = [v for v in lst if v is not None]
        return round(sum(clean) / len(clean), 2) if clean else "—"

    def pct(part, total):
        return f"{round(100 * part / total, 1)}%" if total else "—"

    for i, npc in enumerate(npc_names, start=2):
        rows   = [e for e in results if e["npc"] == npc]
        n      = len(rows)
        gen_ok = sum(1 for e in rows if e["gen_status"] == "OK")
        rv     = sum(1 for e in rows if e.get("rule_valid") is True)
        in_tok  = [e.get("input_tokens")  for e in rows]
        out_tok = [e.get("output_tokens") for e in rows]
        tot_tok = [
            (e.get("input_tokens") or 0) + (e.get("output_tokens") or 0)
            if e.get("input_tokens") is not None and e.get("output_tokens") is not None
            else None
            for e in rows
        ]
        ws.cell(row=i, column=1,  value=npc)
        ws.cell(row=i, column=2,  value=pct(gen_ok, n))
        ws.cell(row=i, column=3,  value=avg([e["gen_latency_s"] for e in rows if e["gen_latency_s"]]))
        ws.cell(row=i, column=4,  value=avg([len(e["steps"]) for e in rows if e["steps"]]))
        ws.cell(row=i, column=5,  value=avg(in_tok))
        ws.cell(row=i, column=6,  value=avg(out_tok))
        ws.cell(row=i, column=7,  value=avg(tot_tok))
        ws.cell(row=i, column=8,  value=pct(rv, n))
        ws.cell(row=i, column=9,  value=avg([e.get("rule_root", 0) for e in rows]))
        ws.cell(row=i, column=10, value=avg([e.get("rule_cascading", 0) for e in rows]))
        _style_row(ws, i, 10, i % 2 == 0)

    # Totals row
    tr         = len(npc_names) + 2
    n_all      = len(results)
    gen_ok_all = sum(1 for e in results if e["gen_status"] == "OK")
    rv_all     = sum(1 for e in results if e.get("rule_valid") is True)
    in_tok_all  = [e.get("input_tokens")  for e in results]
    out_tok_all = [e.get("output_tokens") for e in results]
    tot_tok_all = [
        (e.get("input_tokens") or 0) + (e.get("output_tokens") or 0)
        if e.get("input_tokens") is not None and e.get("output_tokens") is not None
        else None
        for e in results
    ]
    ws.cell(row=tr, column=1,  value="TOTAL / AVERAGE")
    ws.cell(row=tr, column=2,  value=pct(gen_ok_all, n_all))
    ws.cell(row=tr, column=3,  value=avg([e["gen_latency_s"] for e in results if e["gen_latency_s"]]))
    ws.cell(row=tr, column=4,  value=avg([len(e["steps"]) for e in results if e["steps"]]))
    ws.cell(row=tr, column=5,  value=avg(in_tok_all))
    ws.cell(row=tr, column=6,  value=avg(out_tok_all))
    ws.cell(row=tr, column=7,  value=avg(tot_tok_all))
    ws.cell(row=tr, column=8,  value=pct(rv_all, n_all))
    ws.cell(row=tr, column=9,  value=avg([e.get("rule_root", 0) for e in results]))
    ws.cell(row=tr, column=10, value=avg([e.get("rule_cascading", 0) for e in results]))
    _style_row(ws, tr, 10, False)
    for col in range(1, 11):
        ws.cell(row=tr, column=col).font = BOLD_FONT


def save_workbook(scenario_results: dict[str, list[dict]], npc_names: list[str], output_path: str):
    """
    Rebuild the full workbook from all collected results.
    Sheet pairs are written in scenario order: '<Label> Schedules', '<Label> Summary'.
    Called after every NPC so progress is never lost.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for label, results in scenario_results.items():
        if not results:
            continue
        sched_ws = wb.create_sheet(f"{label} Schedules")
        summ_ws  = wb.create_sheet(f"{label} Summary")
        _write_schedules_sheet(sched_ws, results)
        _write_summary_sheet(summ_ws, results, npc_names)

    wb.save(output_path)

# ---------------------------------------------------------------------------
# Per-scenario evaluation
# ---------------------------------------------------------------------------

def run_scenario(label: str, world_state_file: str,
                 npc_names: list[str], characters: list[dict],
                 world_rules: dict, action_library: dict,
                 scenario_results: dict, output_path: str,
                 model: str):
    """Run NUM_RUNS for a single scenario, appending to scenario_results in place."""

    print(f"\n{'#'*60}")
    print(f"SCENARIO: {label}  ({world_state_file})")
    print(f"{'#'*60}")

    world_state = load_json(world_state_file)
    results     = scenario_results.setdefault(label, [])

    for run in range(1, NUM_RUNS + 1):
        print(f"\n  {'='*56}")
        print(f"  RUN {run}/{NUM_RUNS}")
        print(f"  {'='*56}")

        for npc_data in characters:
            npc_name = npc_data.get("name", npc_data.get("id", "unknown"))
            entry = {
                "run":             run,
                "npc":             npc_name,
                "goal":            "",
                "steps":           [],
                "gen_latency_s":   None,
                "input_tokens":    None,
                "output_tokens":   None,
                "gen_status":      "ERROR",
                "rule_valid":      None,
                "rule_violations": [],
                "rule_root":       0,
                "rule_cascading":  0,
            }

            # ── Generate ──────────────────────────────────────────────
            print(f"    [{npc_name}] Generating...", end=" ", flush=True)
            try:
                gen = generate_schedule(npc_name, npc_data, world_state,
                                        world_rules, action_library, model)
                entry.update({
                    "goal":          gen["schedule"]["goal"],
                    "steps":         gen["schedule"]["steps"],
                    "gen_latency_s": gen["latency_s"],
                    "input_tokens":  gen["input_tokens"],
                    "output_tokens": gen["output_tokens"],
                    "gen_status":    "OK",
                })
                tok_str = ""
                if gen["input_tokens"] is not None:
                    tok_str = f", {gen['input_tokens']}in/{gen['output_tokens']}out tokens"
                print(f"OK ({gen['latency_s']}s, {len(gen['schedule']['steps'])} steps{tok_str})")
            except requests.HTTPError as e:
                body = e.response.text if e.response is not None else "(no response body)"
                print(f"FAILED — {e}\n    Server said: {body}")
            except Exception as e:
                print(f"FAILED — {e}")

            # ── Rule-based validation ──────────────────────────────────
            if entry["gen_status"] == "OK":
                report = rule_validate(npc_name, entry["steps"], world_state)
                entry["rule_valid"]      = report.valid
                entry["rule_root"]       = len(report.root_violations)
                entry["rule_cascading"]  = len(report.cascading_violations)
                entry["rule_violations"] = [
                    f"Step {r.step} [{r.action}] ({'CASCADING' if r.cascading else 'ROOT'}): {v}"
                    for r in report.step_results
                    for v in r.violations
                ]
                root   = entry["rule_root"]
                casc   = entry["rule_cascading"]
                status = "VALID" if report.valid else f"INVALID ({root} root, {casc} cascading)"
                print(f"    [{npc_name}] Rule check: {status}")

            results.append(entry)

            # ── Save after every NPC ───────────────────────────────────
            try:
                save_workbook(scenario_results, npc_names, output_path)
            except Exception as e:
                print(f"    [warning] Could not save workbook: {e}")

# ---------------------------------------------------------------------------
# Per-model evaluation
# ---------------------------------------------------------------------------

def run_model(model: str, timestamp: str,
              characters: list[dict], npc_names: list[str],
              world_rules: dict, action_library: dict):
    """Run the full scenario suite for a single model and save its own xlsx."""

    # Sanitise model name for use in filename (replace slashes, spaces, colons)
    safe_model = model.replace("/", "-").replace("\\", "-").replace(" ", "_").replace(":", "-")
    output_path = os.path.join(OUTPUT_DIR, f"npc_evaluation_{timestamp}_{safe_model}.xlsx")

    print(f"\n{'='*60}")
    print(f"MODEL: {model}")
    print(f"Output: {output_path}")
    print(f"{'='*60}")

    # Verify all scenario files exist before starting
    missing = [f for _, f in SCENARIOS if not os.path.exists(os.path.join(DATA_DIR, f))]
    if missing:
        print("ERROR: Missing scenario files:")
        for f in missing:
            print(f"  ../inputs/{f}")
        return

    scenario_results: dict[str, list[dict]] = {}

    for label, world_state_file in SCENARIOS:
        run_scenario(
            label, world_state_file,
            npc_names, characters,
            world_rules, action_library,
            scenario_results, output_path,
            model,
        )

    # Final save
    save_workbook(scenario_results, npc_names, output_path)
    print(f"\nSaved → {output_path}")

    # Per-model summary
    print(f"\n{'='*60}")
    print(f"MODEL COMPLETE: {model}")
    print(f"{'='*60}")
    for label, results in scenario_results.items():
        n        = len(results)
        gen_ok   = sum(1 for e in results if e["gen_status"] == "OK")
        lats     = [e["gen_latency_s"] for e in results if e["gen_latency_s"]]
        avg_lat  = f"{sum(lats)/len(lats):.2f}s" if lats else "—"
        in_toks  = [e["input_tokens"]  for e in results if e["input_tokens"]  is not None]
        out_toks = [e["output_tokens"] for e in results if e["output_tokens"] is not None]
        tok_str  = (
            f"  avg {sum(in_toks)//len(in_toks)}in/{sum(out_toks)//len(out_toks)}out tokens"
            if in_toks else ""
        )
        print(f"  {label:<35} {gen_ok}/{n} OK   avg latency {avg_lat}{tok_str}")

    return scenario_results

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading shared game data...")
    world_rules    = load_json("world_rules.json")
    action_library = load_json("action_library.json")
    characters     = load_json("characters.json")
    npc_names      = [c.get("name", c.get("id", f"npc_{i}")) for i, c in enumerate(characters)]
    print(f"  {len(characters)} characters, "
          f"{len(world_rules.get('rules', world_rules))} rules, "
          f"{len(action_library.get('actions', action_library))} actions\n")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    total_npc_runs = len(MODELS) * len(SCENARIOS) * NUM_RUNS * len(npc_names)
    print(f"Models     : {len(MODELS)}  ({', '.join(MODELS)})")
    print(f"Scenarios  : {len(SCENARIOS)}")
    print(f"Runs each  : {NUM_RUNS}")
    print(f"NPCs each  : {len(npc_names)}")
    print(f"Total calls: {total_npc_runs}")

    for model in MODELS:
        run_model(
            model=model,
            timestamp=timestamp,
            characters=characters,
            npc_names=npc_names,
            world_rules=world_rules,
            action_library=action_library,
        )

    print(f"\n{'#'*60}")
    print("ALL MODELS COMPLETE")
    print(f"{'#'*60}")
    print(f"Output files in: {OUTPUT_DIR}/")
    for model in MODELS:
        safe_model = model.replace("/", "-").replace("\\", "-").replace(" ", "_").replace(":", "-")
        print(f"  npc_evaluation_{timestamp}_{safe_model}.xlsx")


if __name__ == "__main__":
    main()